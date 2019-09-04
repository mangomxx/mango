#!/usr/bin/python
# -*- coding: utf-8 -*-
# @Author  : Miss Mango
import cx_Oracle
import pandas as pd
import os
from openpyxl import load_workbook
file_path = r'C:\Users\18810\Desktop\Python'


def loadDataSet(col,sql):
    """
    加载数据集
    :param fileName: 文件名称
    :return:list类型
    """
    # dataMat = []
    conn = cx_Oracle.connect("BUS_DA/BUS_DA_20181205@172.16.12.121/biwork")
    # 用自己的实际数据库用户名、密码、主机ip地址 替换即可
    curs = conn.cursor()
    curs.execute(sql)
    result = curs.fetchall()
    dataMat = pd.DataFrame(result, columns=col)

    return dataMat



def process_data(df1, key,name):
    # 在每个人的文件夹内，创建个人的图表文件名
    result_file_data = file_path + '/' + key + '.xlsx'
    # 读取key的所有数据，并写入本地文件
    df1.to_excel(result_file_data, sheet_name=name, index=False, encoding="GBK")

    # return df1

# 定义增加sheet表的函数
def excelAddSheet(dataframe, key,name,n):
    outfile = file_path + '/' + key + '.xlsx'
    writer = pd.ExcelWriter(outfile, engine='openpyxl')
    if os.path.exists(outfile) != True:
        dataframe.to_excel(writer, name, index=False)
    else:
        book = load_workbook(writer.path)
        writer.book = book
        dataframe.to_excel(excel_writer=writer, sheet_name=name, startrow=n, index=False)
    writer.save()
    writer.close()


# 定义合并各个明细的函数
def UnionSheet(data1, data2, data3, key, name):
    # 目标文件
    outfile = file_path + '/' + key + '.xlsx'
    writer = pd.ExcelWriter(outfile, engine='openpyxl')
    data= [data1, data2, data3]
    if os.path.exists(outfile) != True:
        data1.to_excel(writer, name, startrow=0, index=False)
        data2.to_excel(writer, name, startrow=int(len(data[0]))+4, index=False)
        # worksheet.write('N15', '总分：')  # Insert an text.
        data3.to_excel(writer, name, startrow=int(len(data[0]))+int(len(data[1]))+10, index=False)
    else:
        book = load_workbook(writer.path)
        writer.book = book
        data1.to_excel(writer, name, startrow=0, index=False)
        data2.to_excel(writer, name, startrow=int(len(data[0])) + 4, index=False)
        data3.to_excel(writer, name, startrow=int(len(data[0])) + int(len(data[1])) + 10, index=False)

        writer.save()
        writer.close()

if __name__ == '__main__':
    col0 = ['name']
    col1 = ['排名', 'name', '区域', '分店', '经纪人', '总业绩', '买卖房源单量', '买卖客户单量', '新增房源量', '新增有效房源量', '活跃房源占比', '钥匙率', '委托率',
            '优图率', '被带看次数', '新增客户量', '新增有效客户量', '带看组数', '二看率', '客均看房量', '成交客户的平均看房套数']
    col2 = ['排名', 'name', '分店', '经纪人', '总业绩', '买卖房源单量', '买卖客户单量', '新增房源量', '新增有效房源量', '活跃房源占比', '钥匙率', '委托率', '优图率',
            '被带看次数', '新增客户量', '新增有效客户量', '带看组数', '二看率', '客均看房量', '成交客户的平均看房套数']
    col3 = ['id', 'name', '楼盘坐落', '带看人', '电话号码', '被带看次数']
    col4 = ['id', 'name', '楼盘坐落', '客户电话', '带看次数']
    col5 = ['客户来源', '比例']
    sql0 = '''SELECT distinct 大区||'-'||区域||'-'||分店||'-'||经纪人
                    FROM BUS_DA.TEST_LEIDA
                    where 大区 in ('西大区')'''  # sql语句取出所有人--左表
    sql1 = '''SELECT  总业绩大区排名,大区,区域,分店,经纪人,总业绩,买卖房源单量,买卖客户单量,新增房源量,新增有效房源量,活跃房源占比,钥匙率,委托率,优图率,被带看次数,新增客户量,新增有效客户量,带看组数,二看率,客均看房量,成交客户的平均看房套数 FROM BUS_DA.CUST_SURVING where 大区 in ('西大区') and 总业绩大区排名<=10 order by 1  '''  # sql语句
    sql2 = '''SELECT  总业绩区域排名,区域,分店,经纪人,总业绩,买卖房源单量,买卖客户单量,新增房源量,新增有效房源量,活跃房源占比,钥匙率,委托率,优图率,被带看次数,新增客户量,新增有效客户量,带看组数,二看率,客均看房量,成交客户的平均看房套数 FROM BUS_DA.CUST_SURVING where 大区 in ('西大区') and 总业绩区域排名<=5 order by 1 '''  # sql语句
    sql3 = '''SELECT 身份证号,大区||'-'||区域||'-'||分店||'-'||经纪人,楼盘坐落,带看人,电话号码,被带看次数 FROM BUS_DA.BEIDAIKAN_SURVING where 大区 in ('西大区')'''  # sql语句
    sql4 = '''SELECT 身份证号,大区||'-'||区域||'-'||分店||'-'||经纪人, 楼盘坐落,客户电话,带看次数 FROM BUS_DA.DAIKAN_SURVING where 大区 in ('西大区') '''  # sql语句
    sql5 = 'SELECT 客户来源,比例 FROM BUS_DA.N_CUST_SOURCE '
    result0 = loadDataSet(col0, sql0)  # 取出key值，exp：西大区-石景山区-远洋山水店A组-张三
    result1 = loadDataSet(col1, sql1)  # 取出大区top10
    result2 = loadDataSet(col2, sql2)  # 取出区域top5
    # 以下三个明细要合并到一张表里
    result3 = loadDataSet(col3, sql3)  # 取出被带看明细
    result4 = loadDataSet(col4, sql4)  # 取出带看明细
    result5 = loadDataSet(col5, sql5)  # 取出客户来源及占比

    id_list = result0['name']  # 取出key值，exp：西大区-石景山区-远洋山水店A组-张三
    # print(id_list)

    for key in id_list:
        # print(key)
        daqu = key.split('-')[0]
        quyu = key.split('-')[1]
        key_data3 = result3.loc[result3['name'] == key]
        key_data4 = result4.loc[result4['name'] == key]
        key_data5 = result5
        key_data1 = result1.loc[result1['name'] == daqu]
        key_data2 = result2.loc[result2['name'] == quyu]
        # excelAddSheet(key_data1, key, '大区Top10', 0)
        # excelAddSheet(key_data2, key, '区域Top5', 0)
        UnionSheet(key_data3, key_data4, key_data5, key, '明细')
    print('finish')



